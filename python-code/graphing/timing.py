import json, time
from time import strftime
import os.path
import os
import pandas as pd
from pandas import ExcelWriter
from pandas import ExcelFile
from openpyxl import load_workbook

def save_time(my_time, type):
    save_date = time.strftime("%m-%d-%Y")
    file_name = './graphing/timing.xlsx'
    excel_file = pd.read_excel(file_name, sheet_name=type, index=0)
    update_flag = True

    for i, row in excel_file.iterrows():
        if excel_file.at[i,'Date'] == save_date:
            print("Previous runtime found for " + type + ". Comparing times...")
            prev_time = excel_file.at[i,'Runtime']
            compare_flag = check_best_time(my_time, prev_time)

            if compare_flag: # Need to update the existing spreadsheet
                print("New best runtime found! Updating spreadsheet...")
                data_frame = [{'Date':save_date,'Runtime':my_time}]
                updated_cell = pd.DataFrame(data_frame)
                append_df_to_excel(file_name, updated_cell, type, i+1)

            update_flag = False
            break
        else:
            continue

    if update_flag: # Add a new entry to the spreadsheet
        print("Adding new runtime for " + type + " on " + save_date + " to spreadsheet...")
        data_frame = [{'Date':save_date,'Runtime':my_time}]
        new_cell = pd.DataFrame(data_frame)
        append_df_to_excel(file_name, new_cell, type, None)

    print("\r")

# Function to compare two times and return a boolean
def check_best_time(new, prev):
    print("New Time: " + str(new) + " vs Previous Time: " + str(prev))
    if(new < prev):
        return True
    else:
        return False

# Function to append a new date to the end of the existing spreadsheet
def append_df_to_excel(filename, df, sheet_name, startrow, truncate_sheet=False, **to_excel_kwargs):
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError

    try:
        writer.book = load_workbook(filename)

        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        if truncate_sheet and sheet_name in writer.book.sheetnames:
            idx = writer.book.sheetnames.index(sheet_name)
            writer.book.remove(writer.book.worksheets[idx])
            writer.book.create_sheet(sheet_name, idx)

        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        pass

    if startrow is None:
        startrow = 0

    df.to_excel(writer, sheet_name, startrow=startrow, header=None, index=None, **to_excel_kwargs)
    writer.save() # write and save
